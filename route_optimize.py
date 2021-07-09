import googlemaps as googlemaps
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
import warnings
import folium
from folium import DivIcon

warnings.filterwarnings("ignore")


class Route:

    def __init__(self, origin, destination):
        self.origin = origin
        self.destination = destination
        self.km = 0
        self.time = 0

    def set_origin(self, new_origin):
        """Simple setter for origin class attribute """
        self.origin = new_origin

    def set_destination(self, new_destination):
        """Simple setter for destination class attribute"""
        self.destination = new_destination

    def get_origin(self):
        """Simple getter for origin class attribute"""
        return self.origin

    def get_destination(self):
        """Simple getter for destination class attribute"""
        return self.destination

    @staticmethod
    def geocode(address, api_key):
        """
        Static method returns geocoded address from string.
        To use it you need to define API as google geocode API

        Parameter:
            > Route.geocode('Plac Defilad 1, 00-901 Warszawa')
            > Route.geocode('Poznań')
        Returns:
            > 52.231918,21.006781
            > 52.406376,16.925167

        Przy nieodnalezieniu adresu zwraca hard-coded adres firmy MPT TABO
        Co skutkować będzie ułożeniem jako pierwszy lub ostatni punkt trasy
        """
        try:
            gmaps = googlemaps.Client(key=api_key)
            geocode_result = gmaps.geocode(address)
            lat = str(geocode_result[0]['geometry']['location']['lat'])
            lng = str(geocode_result[0]['geometry']['location']['lng'])
            return lat + ',' + lng

        except Exception as e:
            print("Błąd przy zamianie adresu: " + str(address) +
                  " na współrzędne geograficzne. " +
                  "Nadałem mu adres firmy.")
            print(e)
            return "52.430540,17.011820"

    @staticmethod
    def create_dataframe_from_excel(excel_file_directory, sheet_name):
        """
        Provided Excel File Directory, which has columns name as
        KOD_KLIENTA|DATA|MIEJSCOWOSC|ULICA|KIER|ORIDEST|KM|KOLEJNOSC|ORIDEST (order is not important)
        where DATA is POSTAL CODE, but another script is running on it, so it had
        to be done this way.

        This will return a file on script directory named: rozpiska_z_kolejnoscia.xlsm
        Having columns as above - KM|KOLEJNOSC will be empty - and its data will be completed
        with other functions. Also there will be added column named latlong for purpose of Google's Directions API.

        Data on this function is easly obtained by pandas functionality.
        From original excel file, at first it deletes empty values on ['KIER']
        Then it is deleting duplicates based on (KOD_P, MIEJSCOWOSC, ULICA, KIEROWCA)
        After that it just saves excel/csv file on script directory.

        parameters:
                excel_file_directory
                sheet_name

        returns:
                dataframe
        """

        try:
            wb = pd.ExcelFile(excel_file_directory)
            sheet1 = pd.read_excel(wb, sheet_name=sheet_name,
                                   usecols=[
                                       'KOD_KLIENTA',
                                       'DATA',
                                       'MIEJSCOWOSC',
                                       'ULICA',
                                       'KIER',
                                       'KOLEJNOSC',
                                       'KM',
                                       'ORIDEST'
                                   ])
            df = sheet1
            df.dropna(subset=['KIER'], inplace=True)
            df.drop_duplicates(subset=['DATA', 'MIEJSCOWOSC', 'ULICA', 'KIER'], inplace=True)
            df['LATLONG'] = ""
            df['FullAddress'] = (df['DATA'].astype(str) +
                                 ' ' + df['MIEJSCOWOSC'].astype(str) +
                                 ' ' + df['ULICA'].astype(str))

            df.to_excel('data.xlsx', index=False)
            return df


        except Exception as e:
            print("Błąd przy odczytywaniu excela")
            print(e)
            exit()

    @staticmethod
    def geocode_dataframe(df, api_key):

        """
        This function takes dataframe and convert cols
        'DATA' + 'MIEJSCOWOSC' + 'ULICA' into LATLONG by geocode function.

        Funkcja uzupełnia kolumnę LATLONG współrzędnymi geograficznymi.

        :param api_key:
        :param df:
        :return: df:
        """

        latlong_index = df.columns.get_loc("LATLONG")

        for i in range(len(df['DATA'])):
            address = (df['FullAddress'].iloc[[i]]).to_string(index=False)
            lat_long = Route.geocode(address, api_key)
            df.iloc[i, latlong_index] = lat_long
            print(str(i) + '/' + str(len(df['DATA'])))
        df.to_excel('data.xlsx', index=False)
        return df

    @staticmethod
    def directions_dataframe(df, api_key):
        """

        parameter
                dataframe
        returns
                dataframe with completed order
        """
        drivers = df
        drivers = drivers['KIER']
        drivers.drop_duplicates(inplace=True)
        drivers = drivers.to_list()
        number_of_drivers = len(drivers)
        nine_hours_from_now = datetime.now() + timedelta(hours=9)
        order_index = df.columns.get_loc("KOLEJNOSC")
        gmaps = googlemaps.Client(key=api_key)

        for i in range(number_of_drivers):
            df_filtered = df.loc[df['KIER'] == drivers[i]]
            waypoints = df_filtered['LATLONG'].array
            """TUTAJ USTAWIA SIE ORIGIN JEŻELI JEST"""
            new_origin = df_filtered.loc[df['ORIDEST'] == 1]
            if new_origin.empty:
                new_origin = ""
            else:
                new_origin = new_origin['FullAddress'].to_string()

            """TUTAJ USTAWIA SIE DESTINATION JEŻELI JEST"""
            new_destination = df_filtered.loc[df['ORIDEST'] == 2]
            if new_destination.empty:
                new_destination = ""
            else:
                new_destination = new_destination['FullAddress'].to_string()

            if new_destination == "" and new_origin == "":
                """no new dest, no new origin"""
                route = Route("52.4356927,17.0300017", "52.4356927,17.0300017")

                try:
                    directions_result = gmaps.directions(route.origin,
                                                         route.destination,
                                                         traffic_model='optimistic',
                                                         departure_time=nine_hours_from_now,
                                                         waypoints=waypoints,
                                                         mode='driving',
                                                         optimize_waypoints=True, )

                    waypoint_order = directions_result[0]['waypoint_order']

                    iteration = 1
                    for order in waypoint_order:
                        # print(str(iteration) + ' ' + df_filtered['KIER'].iloc[order] + ' ' +
                        #       df_filtered['FullAddress'].iloc[order] + ' ' +
                        #       df_filtered['KOD_KLIENTA'].iloc[order])

                        for z in range(len(df['KOD_KLIENTA'])):
                            if (df_filtered['FullAddress'].iloc[order] == df['FullAddress'].iloc[z]
                                    and df_filtered['KIER'].iloc[order] == df['KIER'].iloc[z]):
                                df.iloc[z, order_index] = str(iteration)

                                df.to_excel('data.xlsx', index=False)

                        iteration = iteration + 1
                    df.to_excel('data.xlsx', index=False)
                except Exception as e:
                    print(e)

            if new_destination != "" and new_origin == "":
                """there is destination, there isn't origin"""
                new_destination = Route.geocode(new_destination, api_key)
                route = Route("52.4356927,17.0300017", new_destination)
                try:
                    directions_result = gmaps.directions(route.origin,
                                                         route.destination,
                                                         traffic_model='optimistic',
                                                         departure_time=nine_hours_from_now,
                                                         waypoints=waypoints,
                                                         mode='driving',
                                                         optimize_waypoints=True, )

                    waypoint_order = directions_result[0]['waypoint_order']

                    iteration = 1
                    for order in waypoint_order:
                        # print(str(iteration) + ' ' + df_filtered['KIER'].iloc[order] + ' ' +
                        #       df_filtered['FullAddress'].iloc[order] + ' ' +
                        #       df_filtered['KOD_KLIENTA'].iloc[order])

                        for z in range(len(df['KOD_KLIENTA'])):
                            if (df_filtered['FullAddress'].iloc[order] == df['FullAddress'].iloc[z]
                                    and df_filtered['KIER'].iloc[order] == df['KIER'].iloc[z]):
                                df.iloc[z, order_index] = str(iteration)

                                df.to_excel('data.xlsx', index=False)

                        iteration = iteration + 1
                    df.to_excel('data.xlsx', index=False)
                except Exception as e:
                    print(e)

            if new_destination == "" and new_origin != "":
                """There is new origin"""
                new_origin = Route.geocode(new_origin, api_key)
                route = Route(new_origin, "52.4356927,17.0300017")
                try:
                    directions_result = gmaps.directions(route.origin,
                                                         route.destination,
                                                         traffic_model='optimistic',
                                                         departure_time=nine_hours_from_now,
                                                         waypoints=waypoints,
                                                         mode='driving',
                                                         optimize_waypoints=True, )

                    waypoint_order = directions_result[0]['waypoint_order']

                    iteration = 1
                    for order in waypoint_order:
                        # print(str(iteration) + ' ' + df_filtered['KIER'].iloc[order] + ' ' +
                        #       df_filtered['FullAddress'].iloc[order] + ' ' +
                        #       df_filtered['KOD_KLIENTA'].iloc[order])

                        for z in range(len(df['KOD_KLIENTA'])):
                            if (df_filtered['FullAddress'].iloc[order] == df['FullAddress'].iloc[z]
                                    and df_filtered['KIER'].iloc[order] == df['KIER'].iloc[z]):
                                df.iloc[z, order_index] = str(iteration)

                                df.to_excel('data.xlsx', index=False)

                        iteration = iteration + 1
                    df.to_excel('data.xlsx', index=False)
                except Exception as e:
                    print(e)

            if new_destination != "" and new_origin != "":

                """ Case with new dest and origin """

                new_destination = Route.geocode(new_destination, api_key)
                new_origin = Route.geocode(new_origin, api_key)

                route = Route(new_origin, new_destination)

                try:

                    directions_result = gmaps.directions(route.origin,
                                                         route.destination,
                                                         traffic_model='optimistic',
                                                         departure_time=nine_hours_from_now,
                                                         waypoints=waypoints,
                                                         mode='driving',
                                                         optimize_waypoints=True, )

                    waypoint_order = directions_result[0]['waypoint_order']

                    iteration = 1
                    for order in waypoint_order:
                        # print(str(iteration) + ' ' + df_filtered['KIER'].iloc[order] + ' ' +
                        #       df_filtered['FullAddress'].iloc[order] + ' ' +
                        #       df_filtered['KOD_KLIENTA'].iloc[order])

                        for z in range(len(df['KOD_KLIENTA'])):
                            if (df_filtered['FullAddress'].iloc[order] == df['FullAddress'].iloc[z]
                                    and df_filtered['KIER'].iloc[order] == df['KIER'].iloc[z]):
                                df.iloc[z, order_index] = str(iteration)

                                df.to_excel('data.xlsx', index=False)

                        iteration = iteration + 1
                    df.to_excel('data.xlsx', index=False)
                except Exception as e:
                    print(e)

        return df

    @staticmethod
    def complete_km(df, api_key):
        drivers = df
        drivers = drivers['KIER']
        drivers.drop_duplicates(inplace=True)
        drivers = drivers.to_list()
        number_of_drivers = len(drivers)
        nine_hours_from_now = datetime.now() + timedelta(hours=9)
        km_index = df.columns.get_loc("KM")
        gmaps = googlemaps.Client(key=api_key)

        kilometry = []

        for i in range(number_of_drivers):
            df_filtered = df.loc[df['KIER'] == drivers[i]]
            waypoints = df_filtered['LATLONG'].array

            route = Route("52.4356927,17.0300017", "52.4356927,17.0300017")

            try:
                directions_result = gmaps.directions(route.origin,
                                                     route.destination,
                                                     traffic_model='optimistic',
                                                     departure_time=nine_hours_from_now,
                                                     waypoints=waypoints,
                                                     mode='driving',
                                                     optimize_waypoints=True, )
                distance = 0
                time = 0
                legs = directions_result[0].get("legs")
                for leg in legs:
                    distance = distance + leg.get("distance").get("value")
                    time = time + leg.get("duration").get("value")

                print('\nKIEROWCA: ', drivers[i])
                print('Total distance: ', distance / 1000, ' km')
                time = time % (24 * 3600)
                hour = time // 3600
                time %= 3600
                minutes = time // 60
                time %= 60
                print("Czas jazdy: %d h %d m" % (hour, minutes))
                kilometry.append(i)
                kilometry.append(distance / 1000)

                waypoint_order = directions_result[0]['waypoint_order']
                iteration = 1
                for order in waypoint_order:
                    for z in range(len(df['KOD_KLIENTA'])):
                        if (df_filtered['FullAddress'].iloc[order] == df['FullAddress'].iloc[z]
                                and df_filtered['KIER'].iloc[order] == df['KIER'].iloc[z]):
                            df.iloc[z, km_index] = str(distance / 1000)

                            df.to_excel('data.xlsx', index=False)

                    iteration = iteration + 1
                    df.to_excel('data.xlsx', index=False)
            except Exception as e:
                print(e)
            df.to_excel('data.xlsx', index=False)

        return df

    @staticmethod
    def connect_excel_with_df(df, file_directory, sheet_name):
        # wyjmuje index z dataframe
        wb = pd.ExcelFile(file_directory)
        sheet1 = pd.read_excel(wb, sheet_name=sheet_name)
        data_frame = sheet1


        kod_klienta_index = data_frame.columns.get_loc("KOD_KLIENTA")
        kod_pocztowy_index = data_frame.columns.get_loc("DATA")
        miejscowosc_index = data_frame.columns.get_loc("MIEJSCOWOSC")
        ulica_index = data_frame.columns.get_loc("ULICA")
        kier_index = data_frame.columns.get_loc("KIER")
        kolejnosc_index = data_frame.columns.get_loc("KOLEJNOSC")
        km_index = data_frame.columns.get_loc("KM")




        wb = openpyxl.load_workbook(file_directory)
        sheet = wb.get_sheet_by_name(sheet_name)


        iteracja = 1

        while iteracja < 3500:

            kod_klienta = sheet.cell(row=iteracja, column=kod_klienta_index+1)
            kod_pocztowy = sheet.cell(row=iteracja, column=kod_pocztowy_index+1)
            miejscowosc = sheet.cell(row=iteracja, column=miejscowosc_index+1)
            ulica = sheet.cell(row=iteracja, column=ulica_index+1)
            kier = sheet.cell(row=iteracja, column=kier_index+1)
            kolejnosc = sheet.cell(row=iteracja, column=kolejnosc_index+1)
            kilometry = sheet.cell(row=iteracja, column=km_index+1)

            if kod_klienta.value is None:
                iteracja = iteracja + 1
                continue
            else:
                for z in range(len(df["KOD_KLIENTA"])):

                    if (kod_pocztowy.value == df['DATA'].iloc[z] and
                            miejscowosc.value == df['MIEJSCOWOSC'].iloc[z] and
                            ulica.value == df['ULICA'].iloc[z] and
                            kier.value == df['KIER'].iloc[z]):
                        kolejnosc.value = df['KOLEJNOSC'].iloc[z]
                        kilometry.value = df['KM'].iloc[z]

            iteracja = iteracja + 1
        wb.save('rozpiska_z_kolejnoscia.xlsx')
        wb.close()
        return df

    @staticmethod
    def create_map():

        wb = pd.ExcelFile('data.xlsx')
        sheet1 = None
        try:
            sheet1 = pd.read_excel(wb, sheet_name='Sheet1')
        except Exception as e:
            print(e)
            exit()

        gk = sheet1
        jaki_kierowca = 'continue'

        wybor = input("Chcesz mape? Y/N ")
        if str(wybor) != "Y":
            exit()



        while jaki_kierowca != 'exit':

            jaki_kierowca = input('Jakiego kierowce chcesz podejrzeć? ')
            m = folium.Map(location=[52.430538, 17.011820], control_scale=False)

            zk = gk
            zk = zk[zk['KIER'] == jaki_kierowca]
            dlugosc = len(zk['KOD_KLIENTA'])
            zk = zk.sort_values(by='KOLEJNOSC')
            print(zk)

            iteration = 1

            for x in range(dlugosc):

                lat, lng = zk.iloc[x]['LATLONG'].split(',')

                folium.Marker([lat, lng], popup=zk.iloc[x]['KOD_KLIENTA'],
                              icon_size=(14, 14),
                              show=True, tooltip=zk.iloc[x]['FullAddress'], permanent=True,
                              icon=DivIcon(
                                  icon_size=(14, 14),
                                  icon_anchor=(0, 0),
                                  html='<div style="font-size: 8pt"><b>' +
                                       str(zk.iloc[x]['KOLEJNOSC']) + '</div>',
                              )).add_to(m)

                folium.Marker([lat, lng], popup=zk.iloc[x]['KOD_KLIENTA'], show=True, tooltip=zk.iloc[x]['FullAddress'],
                              permanent=True).add_to(m)

                iteration = iteration + 1
            m.save('mapa.html')
